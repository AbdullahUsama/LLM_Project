"""
Extract Q&A pairs from NUST Bank-Product-Knowledge.xlsx and format
them for LLM fine-tuning.

Outputs:
  1. data/finetuning_data.jsonl       – one JSON object per line (instruction / input / output)
  2. data/finetuning_data_chat.jsonl   – OpenAI chat-format (messages array per line)
  3. data/all_qa_pairs.json            – human-readable dump of every extracted Q&A

The Excel is NOT a clean tabular dataset – it stores Q&A knowledge in
merged cells across 34+ sheets, each with a slightly different layout.
This script handles:
  • Merged cell ranges (openpyxl)
  • Multi-row answers that need to be collapsed into a single string
  • Side-by-side columns (some sheets put extra info in cols F-N)
  • Rate Sheet / Main index / empty sheets → skipped
"""

import json
import os
import re
import sys
import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
XLSX_PATH   = os.path.join(PROJECT_DIR, "NUST Bank-Product-Knowledge.xlsx")

if PROJECT_DIR not in sys.path:
    sys.path.insert(0, PROJECT_DIR)

from src.data_pipeline import ingest_new_qa_pairs

OUT_JSONL_INSTRUCT = os.path.join(SCRIPT_DIR, "finetuning_data.jsonl")
OUT_JSONL_CHAT     = os.path.join(SCRIPT_DIR, "finetuning_data_chat.jsonl")
OUT_JSON_ALL       = os.path.join(SCRIPT_DIR, "all_qa_pairs.json")

# Sheets to skip (index page, rate sheet, empty)
SKIP_SHEETS = {"Main", "Rate Sheet July 1 2024", "Sheet1"}

# Product-friendly names derived from sheet row-1 titles
# (we read them dynamically from the first row of each sheet)

SYSTEM_PROMPT = (
    """System:
You are a helpful customer support assistant for NUST Bank.

Your job is to answer questions about NUST Bank products, services, account features, eligibility, fees, limits, procedures, mobile banking, internet banking, transfers, and related FAQs.

Rules:
- Answer only from the provided context.
- If the question is outside NUST Bank product or app scope, say you can only help with NUST Bank product and app questions.
- If the context does not contain enough information, say: "I don't have enough information in the provided knowledge base to answer that accurately."
- Do not guess, infer, or invent details.
- Do not add policies, fees, limits, eligibility rules, or procedures unless they are explicitly in the context.
- Preserve exact product names, numbers, percentages, limits, dates, and conditions from the context.
- If multiple context chunks conflict, mention the conflict and avoid guessing.
- Keep the answer concise, factual, and directly responsive.
- If the question is ambiguous, ask one short clarifying question.
"""
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    """Normalise whitespace, strip, and collapse blank lines."""
    if not text:
        return ""
    text = text.strip()
    # Replace tabs with spaces
    text = text.replace("\t", " ")
    # Collapse multiple spaces (but keep newlines meaningful)
    text = re.sub(r"[^\S\n]+", " ", text)
    # Collapse 3+ newlines into 2
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def is_question(text: str) -> bool:
    """Heuristic: does this text look like a question?"""
    t = text.strip()
    if not t:
        return False
    # Ends with question mark
    if t.endswith("?"):
        return True
    # Starts with common question patterns
    q_starts = (
        "what", "how", "is ", "is\n", "can ", "can\n", "does", "do ",
        "are ", "which", "who", "where", "when", "why",
        "i would like to", "i want to", "please tell",
        "1.", "1 .", "1-",  # numbered questions like "1.What are..."
    )
    lower = t.lower()
    for qs in q_starts:
        if lower.startswith(qs):
            return True
    # Contains a question mark anywhere (sometimes merged text)
    if "?" in t[:80]:
        return True
    return False


def get_all_cell_values(ws):
    """
    Return a list of (row_idx, col_letter, value_str) for every non-empty,
    non-formula cell in the worksheet.
    """
    results = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            # Skip Excel formula references
            if val.startswith("="):
                continue
            if val.lower() == "main":
                continue
            if not val:
                continue
            results.append((cell.row, cell.column_letter, val))
    return results


def extract_qa_from_sheet(ws, sheet_name: str) -> list[dict]:
    """
    Walk through a sheet, detect question rows and their answers.
    Returns a list of {"question": ..., "answer": ..., "product": ...}
    """
    # Get product name from first row
    product_name = sheet_name  # fallback
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        for cell in row:
            if cell.value and not str(cell.value).strip().lower() == "main":
                val = str(cell.value).strip()
                if not val.startswith("="):
                    product_name = val
                    break

    # Collect all non-empty text cells, grouped by row
    rows_data = {}  # row_num -> list of text values
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        row_texts = []
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if val.startswith("=") or val.lower() == "main" or not val:
                continue
            row_texts.append(val)
        if row_texts:
            # Combine all cell values in the row into one string
            combined = " | ".join(row_texts) if len(row_texts) > 1 else row_texts[0]
            rows_data[row.row if hasattr(row, 'row') else row[0].row] = combined

    if not rows_data:
        return []

    # Sort by row number
    sorted_rows = sorted(rows_data.items())

    # Skip the first row if it's just the product title
    start_idx = 0
    if sorted_rows and sorted_rows[0][1].strip() == product_name.strip():
        start_idx = 1

    # Now walk through rows detecting Q&A pairs
    qa_pairs = []
    current_question = None
    current_answer_parts = []

    for row_num, text in sorted_rows[start_idx:]:
        text = clean_text(text)
        if not text:
            continue

        if is_question(text):
            # Save previous Q&A if exists
            if current_question and current_answer_parts:
                answer = "\n".join(current_answer_parts)
                answer = clean_text(answer)
                if answer:
                    qa_pairs.append({
                        "question": clean_text(current_question),
                        "answer": answer,
                        "product": product_name,
                        "sheet": sheet_name,
                    })
            elif current_question and not current_answer_parts:
                # Question with no answer — skip or note
                pass

            current_question = text
            current_answer_parts = []
        else:
            # This is an answer line (or continuation)
            if current_question:
                current_answer_parts.append(text)
            else:
                # Answer before any question — could be intro text, skip
                pass

    # Don't forget the last Q&A
    if current_question and current_answer_parts:
        answer = "\n".join(current_answer_parts)
        answer = clean_text(answer)
        if answer:
            qa_pairs.append({
                "question": clean_text(current_question),
                "answer": answer,
                "product": product_name,
                "sheet": sheet_name,
            })

    return qa_pairs


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    all_qa = []
    skipped = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            skipped.append(sheet_name)
            continue

        ws = wb[sheet_name]
        pairs = extract_qa_from_sheet(ws, sheet_name)
        all_qa.extend(pairs)
        print(f"  ✓ {sheet_name:25s} → {len(pairs):3d} Q&A pairs")

    print(f"\nSkipped sheets: {skipped}")
    print(f"Total Q&A pairs extracted: {len(all_qa)}")

    # Also load the existing JSON FAQ if present
    json_faq_path = os.path.join(PROJECT_DIR, "funds_transer_app_features_faq.json")
    if os.path.exists(json_faq_path):
        with open(json_faq_path, "r", encoding="utf-8") as f:
            faq_data = json.load(f)
        for cat in faq_data.get("categories", []):
            cat_name = cat.get("category", "General")
            for q in cat.get("questions", []):
                all_qa.append({
                    "question": clean_text(q["question"]),
                    "answer": clean_text(q["answer"]),
                    "product": f"Mobile App – {cat_name}",
                    "sheet": "funds_transfer_app_features_faq.json",
                })
        print(f"  ✓ {'JSON FAQ':25s} → {sum(len(c.get('questions', [])) for c in faq_data.get('categories', [])):3d} Q&A pairs")
        print(f"Total Q&A pairs (with JSON): {len(all_qa)}")

    # ── Append-safe write/update pipeline ──────────────────────────────────────
    summary = ingest_new_qa_pairs(all_qa)

    print(f"\n✅ Updated all Q&A pairs      → {OUT_JSON_ALL}")
    print(f"✅ Updated instruction JSONL  → {OUT_JSONL_INSTRUCT}")
    print(f"✅ Updated chat JSONL         → {OUT_JSONL_CHAT}")

    # ── Summary stats ─────────────────────────────────────────────────────────
    products = {}
    for qa in all_qa:
        products.setdefault(qa["product"], 0)
        products[qa["product"]] += 1

    print(f"\n{'─'*60}")
    print(f"{'PRODUCT':<45} {'Q&A COUNT':>10}")
    print(f"{'─'*60}")
    for prod, count in sorted(products.items(), key=lambda x: -x[1]):
        print(f"  {prod:<43} {count:>10}")
    print(f"{'─'*60}")
    print(f"  {'TOTAL':<43} {len(all_qa):>10}")

    print(f"\nAppend summary: {summary}")

    # Quick quality check – show a few samples
    print(f"\n{'═'*60}")
    print("SAMPLE Q&A PAIRS (first 3):")
    print(f"{'═'*60}")
    for i, qa in enumerate(all_qa[:3]):
        print(f"\n  [{i+1}] Product: {qa['product']}")
        print(f"      Q: {qa['question'][:120]}")
        print(f"      A: {qa['answer'][:200]}")


if __name__ == "__main__":
    main()

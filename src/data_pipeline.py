from __future__ import annotations

import json
import io
import os
import re
from pathlib import Path
from typing import Iterable

import faiss
import numpy as np
import openpyxl
from langchain_community.vectorstores import FAISS as LangChainFAISS
from langchain_core.documents import Document
from langchain_huggingface import HuggingFaceEmbeddings
from sentence_transformers import SentenceTransformer

SRC_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SRC_DIR.parent

# Prefer the repository-level data directory so ingestion and retrieval stay in sync.
if (PROJECT_ROOT / "data").exists():
    DATA_DIR = PROJECT_ROOT / "data"
else:
    # Backward-compatible fallback for older layouts.
    DATA_DIR = SRC_DIR / "data"
EMBEDDING_MODEL = "all-MiniLM-L6-v2"
ANONYMIZE_ON_INGEST = os.getenv("ANONYMIZE_ON_INGEST", "true").strip().lower() not in {
    "0",
    "false",
    "no",
    "off",
}
SYSTEM_PROMPT = (
    """System:
You are a helpful customer support assistant for NUST Bank.

Your job is to answer questions about NUST Bank products, services, account features, eligibility, fees, limits, procedures, mobile banking, internet banking, transfers, and related FAQs.

Rules:
- Answer only from the provided context.
- If the question is outside NUST Bank product or app scope, say you can only help with NUST Bank product and app questions.
- If the context does not contain enough information, say: "I don't have enough information in the provided knowledge base to answer that accurately."
- Do not guess, infer, or invent details.
- Do not add policies, fees, limits, eligibility rules, or procedures unless they are explicitly in the context.
- Preserve exact product names, numbers, percentages, limits, dates, and conditions from the context.
- If multiple context chunks conflict, mention the conflict and avoid guessing.
- Keep the answer concise, factual, and directly responsive.
- If the question is ambiguous, ask one short clarifying question.
"""
)

ALL_QA_PATH = DATA_DIR / "all_qa_pairs.json"
FINETUNE_INSTRUCT_PATH = DATA_DIR / "finetuning_data.jsonl"
FINETUNE_CHAT_PATH = DATA_DIR / "finetuning_data_chat.jsonl"
CHUNK_METADATA_PATH = DATA_DIR / "chunk_metadata.json"
RAW_FAISS_INDEX_PATH = DATA_DIR / "faiss_index.bin"
VECTORSTORE_DIR = DATA_DIR / "vectorstore"


def clean_text(text: str) -> str:
    if not text:
        return ""
    text = text.strip().replace("\t", " ")
    text = re.sub(r"[^\S\n]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def anonymize_text(text: str) -> str:
    if not text:
        return ""

    sanitized = clean_text(text)
    if not ANONYMIZE_ON_INGEST:
        return sanitized

    replacements = [
        (r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", "[EMAIL]"),
        (
            r"\b(?:\+?92[-\s]?)?(?:0)?3\d{2}[-\s]?\d{7}\b|\b\d{3,5}[-\s]?\d{6,8}\b",
            "[PHONE]",
        ),
        (r"\b\d{5}-?\d{7}-?\d\b", "[CNIC]"),
        (r"\b\d{13}\b", "[CNIC]"),
        (r"\b(?:\d[ -]*?){13,19}\b", "[CARD_NUMBER]"),
        (r"(?i)\b(?:iban|account|a/c|acc(?:ount)?)\b(?:\s*(?:no\.?|number|#)?\s*[:\-]?\s*\d{6,18})", "[ACCOUNT_NUMBER]"),
        (r"\b[A-Z]{2}\d{2}[A-Z0-9]{10,30}\b", "[IBAN]"),
    ]

    for pattern, replacement in replacements:
        sanitized = re.sub(pattern, replacement, sanitized)

    return sanitized


def normalize_text(text: str) -> str:
    text = re.sub(r"\s*\|\s*", " ", text)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"^[•◦\-\d\.]+\s*", "", text, flags=re.MULTILINE)
    return text.strip()


def is_question(text: str) -> bool:
    t = text.strip()
    if not t:
        return False
    if t.endswith("?"):
        return True
    q_starts = (
        "what", "how", "is ", "is\n", "can ", "can\n", "does", "do ",
        "are ", "which", "who", "where", "when", "why",
        "i would like to", "i want to", "please tell",
        "1.", "1 .", "1-",
    )
    lower = t.lower()
    for qs in q_starts:
        if lower.startswith(qs):
            return True
    if "?" in t[:80]:
        return True
    return False


def extract_qa_from_sheet(ws, sheet_name: str) -> list[dict]:
    product_name = sheet_name
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        for cell in row:
            if cell.value and not str(cell.value).strip().lower() == "main":
                val = str(cell.value).strip()
                if not val.startswith("="):
                    product_name = val
                    break

    # First try tabular extraction (Question/Answer columns), which is common in FAQ workbooks.
    header_row_idx = None
    question_col_idx = None
    answer_col_idx = None
    product_col_idx = None

    scan_limit = min(ws.max_row, 15)
    for row_idx in range(1, scan_limit + 1):
        row_values = [str(v).strip().lower() if v is not None else "" for v in ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True).__next__()]
        for col_idx, value in enumerate(row_values):
            if question_col_idx is None and value in {"question", "questions", "query", "faq question"}:
                question_col_idx = col_idx
            if answer_col_idx is None and value in {"answer", "answers", "response", "faq answer"}:
                answer_col_idx = col_idx
            if product_col_idx is None and value in {"product", "category", "service", "segment"}:
                product_col_idx = col_idx

        if question_col_idx is not None and answer_col_idx is not None:
            header_row_idx = row_idx
            break

    if header_row_idx is not None:
        tabular_pairs: list[dict] = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
            q_val = row[question_col_idx] if question_col_idx < len(row) else None
            a_val = row[answer_col_idx] if answer_col_idx < len(row) else None
            question = clean_text(str(q_val)) if q_val is not None else ""
            answer = clean_text(str(a_val)) if a_val is not None else ""
            if not question or not answer:
                continue

            row_product = product_name
            if product_col_idx is not None and product_col_idx < len(row):
                p_val = row[product_col_idx]
                if p_val is not None and clean_text(str(p_val)):
                    row_product = clean_text(str(p_val))

            tabular_pairs.append(
                {
                    "question": question,
                    "answer": answer,
                    "product": row_product,
                    "sheet": sheet_name,
                }
            )

        if tabular_pairs:
            return tabular_pairs

    rows_data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        row_texts = []
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if val.startswith("=") or val.lower() == "main" or not val:
                continue
            row_texts.append(val)
        if row_texts:
            combined = " | ".join(row_texts) if len(row_texts) > 1 else row_texts[0]
            rows_data[row[0].row] = combined

    if not rows_data:
        return []

    sorted_rows = sorted(rows_data.items())
    start_idx = 0
    if sorted_rows and sorted_rows[0][1].strip() == product_name.strip():
        start_idx = 1

    qa_pairs = []
    current_question = None
    current_answer_parts = []

    for _, text in sorted_rows[start_idx:]:
        text = clean_text(text)
        if not text:
            continue

        if is_question(text):
            if current_question and current_answer_parts:
                answer = clean_text("\n".join(current_answer_parts))
                if answer:
                    qa_pairs.append({
                        "question": clean_text(current_question),
                        "answer": answer,
                        "product": product_name,
                        "sheet": sheet_name,
                    })
            current_question = text
            current_answer_parts = []
        elif current_question:
            current_answer_parts.append(text)

    if current_question and current_answer_parts:
        answer = clean_text("\n".join(current_answer_parts))
        if answer:
            qa_pairs.append({
                "question": clean_text(current_question),
                "answer": answer,
                "product": product_name,
                "sheet": sheet_name,
            })

    return qa_pairs


def extract_qa_pairs_from_workbook(workbook_bytes: bytes, filename: str = "uploaded.xlsx") -> list[dict]:
    skip_sheets = {"Main", "Rate Sheet July 1 2024", "Sheet1"}
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    all_pairs = []
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        all_pairs.extend(extract_qa_from_sheet(ws, sheet_name))
    return all_pairs


def load_json_list(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, list) else []


def save_json_list(path: Path, data: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def append_jsonl_records(path: Path, records: Iterable[dict]) -> int:
    records = list(records)
    if not records:
        return 0
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    return len(records)


def canonicalize_qa_pair(
    item: dict,
    default_product: str = "Manual Entry",
    default_sheet: str = "Manual Input",
) -> dict | None:
    question = anonymize_text(item.get("question", ""))
    answer = anonymize_text(item.get("answer", ""))
    if not question or not answer:
        return None
    return {
        "question": question,
        "answer": answer,
        "product": clean_text(item.get("product") or default_product) or default_product,
        "sheet": clean_text(item.get("sheet") or default_sheet) or default_sheet,
    }


def qa_signature(item: dict) -> str:
    return " || ".join(
        [
            normalize_text(item.get("question", "")).lower(),
            normalize_text(item.get("answer", "")).lower(),
            normalize_text(item.get("product", "")).lower(),
            normalize_text(item.get("sheet", "")).lower(),
        ]
    )


def build_chunks_from_qa_pairs(pairs: list[dict], start_id: int = 0) -> list[dict]:
    chunks = []
    for offset, item in enumerate(pairs):
        question = normalize_text(item["question"])
        answer = normalize_text(item["answer"])
        chunk = {
            "id": start_id + offset,
            "text": f"Question: {question}\nAnswer: {answer}",
            "question": question,
            "answer": answer,
            "product": item.get("product", "Unknown"),
            "sheet": item.get("sheet", "Unknown"),
        }
        chunks.append(chunk)
    return chunks


def append_training_files(new_pairs: Iterable[dict]) -> dict:
    new_pairs = list(new_pairs)
    existing_pairs = load_json_list(ALL_QA_PATH)
    existing_signatures = {qa_signature(item) for item in existing_pairs}

    appended_pairs: list[dict] = []
    for item in new_pairs:
        canonical = canonicalize_qa_pair(item)
        if not canonical:
            continue
        signature = qa_signature(canonical)
        if signature in existing_signatures:
            continue
        existing_signatures.add(signature)
        appended_pairs.append(canonical)

    if not appended_pairs:
        return {
            "added": 0,
            "skipped_duplicates": len(new_pairs),
            "total_all_qa": len(existing_pairs),
        }

    updated_pairs = existing_pairs + appended_pairs
    save_json_list(ALL_QA_PATH, updated_pairs)

    append_jsonl_records(
        FINETUNE_INSTRUCT_PATH,
        (
            {
                "instruction": pair["question"],
                "input": "",
                "output": pair["answer"],
                "product": pair["product"],
            }
            for pair in appended_pairs
        ),
    )
    append_jsonl_records(
        FINETUNE_CHAT_PATH,
        (
            {
                "messages": [
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": pair["question"]},
                    {"role": "assistant", "content": pair["answer"]},
                ]
            }
            for pair in appended_pairs
        ),
    )

    return {
        "added": len(appended_pairs),
        "skipped_duplicates": len(new_pairs) - len(appended_pairs),
        "total_all_qa": len(updated_pairs),
    }


def append_raw_faiss_index() -> dict:
    all_pairs = load_json_list(ALL_QA_PATH)
    existing_chunks = load_json_list(CHUNK_METADATA_PATH)

    if RAW_FAISS_INDEX_PATH.exists() and existing_chunks:
        index = faiss.read_index(str(RAW_FAISS_INDEX_PATH))
        existing_count = min(index.ntotal, len(existing_chunks))
        if existing_count < len(all_pairs):
            new_pairs = all_pairs[existing_count:]
            new_chunks = build_chunks_from_qa_pairs(new_pairs, start_id=existing_count)
            model = SentenceTransformer(EMBEDDING_MODEL)
            embeddings = model.encode(
                [chunk["text"] for chunk in new_chunks],
                show_progress_bar=True,
                convert_to_numpy=True,
            ).astype(np.float32)
            index.add(embeddings)
            faiss.write_index(index, str(RAW_FAISS_INDEX_PATH))
            save_json_list(CHUNK_METADATA_PATH, existing_chunks[:existing_count] + new_chunks)
            return {
                "added_chunks": len(new_chunks),
                "total_chunks": existing_count + len(new_chunks),
                "index_size": index.ntotal,
            }
        return {
            "added_chunks": 0,
            "total_chunks": existing_count,
            "index_size": index.ntotal,
        }

    # Recovery path: rebuild from the current source data if the raw index is missing or out of sync.
    if not all_pairs:
        return {"added_chunks": 0, "total_chunks": 0, "index_size": 0}

    chunks = build_chunks_from_qa_pairs(all_pairs, start_id=0)
    model = SentenceTransformer(EMBEDDING_MODEL)
    embeddings = model.encode(
        [chunk["text"] for chunk in chunks],
        show_progress_bar=True,
        convert_to_numpy=True,
    ).astype(np.float32)
    index = faiss.IndexFlatL2(embeddings.shape[1])
    index.add(embeddings)

    RAW_FAISS_INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    faiss.write_index(index, str(RAW_FAISS_INDEX_PATH))
    save_json_list(CHUNK_METADATA_PATH, chunks)

    return {"added_chunks": len(chunks), "total_chunks": len(chunks), "index_size": index.ntotal}


def append_langchain_vectorstore() -> dict:
    chunks = load_json_list(CHUNK_METADATA_PATH)
    if not chunks:
        return {"added_documents": 0, "total_documents": 0}

    embeddings = HuggingFaceEmbeddings(model_name=EMBEDDING_MODEL)
    VECTORSTORE_DIR.mkdir(parents=True, exist_ok=True)

    if (VECTORSTORE_DIR / "index.faiss").exists() and (VECTORSTORE_DIR / "index.pkl").exists():
        vectorstore = LangChainFAISS.load_local(
            str(VECTORSTORE_DIR),
            embeddings,
            allow_dangerous_deserialization=True,
        )
        existing_count = vectorstore.index.ntotal
        if existing_count > len(chunks):
            documents = [
                Document(
                    page_content=chunk["text"],
                    metadata={
                        "id": chunk["id"],
                        "question": chunk["question"],
                        "answer": chunk["answer"],
                        "product": chunk.get("product", "Unknown"),
                        "sheet": chunk.get("sheet", "Unknown"),
                    },
                )
                for chunk in chunks
            ]
            vectorstore = LangChainFAISS.from_documents(documents, embeddings)
            vectorstore.save_local(str(VECTORSTORE_DIR))
            return {"added_documents": len(documents), "total_documents": vectorstore.index.ntotal}

        new_chunks = chunks[existing_count:]
        if not new_chunks:
            return {"added_documents": 0, "total_documents": existing_count}

        documents = [
            Document(
                page_content=chunk["text"],
                metadata={
                    "id": chunk["id"],
                    "question": chunk["question"],
                    "answer": chunk["answer"],
                    "product": chunk.get("product", "Unknown"),
                    "sheet": chunk.get("sheet", "Unknown"),
                },
            )
            for chunk in new_chunks
        ]
        vectorstore.add_documents(documents)
        vectorstore.save_local(str(VECTORSTORE_DIR))
        return {
            "added_documents": len(new_chunks),
            "total_documents": vectorstore.index.ntotal,
        }

    documents = [
        Document(
            page_content=chunk["text"],
            metadata={
                "id": chunk["id"],
                "question": chunk["question"],
                "answer": chunk["answer"],
                "product": chunk.get("product", "Unknown"),
                "sheet": chunk.get("sheet", "Unknown"),
            },
        )
        for chunk in chunks
    ]
    vectorstore = LangChainFAISS.from_documents(documents, embeddings)
    vectorstore.save_local(str(VECTORSTORE_DIR))
    return {"added_documents": len(documents), "total_documents": vectorstore.index.ntotal}


def ingest_new_qa_pairs(new_pairs: Iterable[dict]) -> dict:
    new_pairs = list(new_pairs)
    training_summary = append_training_files(new_pairs)
    raw_summary = append_raw_faiss_index()
    vector_summary = append_langchain_vectorstore()
    return {
        "training": training_summary,
        "raw_faiss": raw_summary,
        "vectorstore": vector_summary,
    }
